from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

wb = Workbook()

# ── Colour palette ──
NAVY = "1B2A4A"
DARK_BLUE = "2C3E6B"
MID_BLUE = "3A5BA0"
LIGHT_BLUE = "D6E4F0"
PALE_BLUE = "EBF1FA"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"
MID_GREY = "D9D9D9"
DARK_GREY = "404040"
AMBER = "F4A460"
GREEN_BG = "E2EFDA"
RED_BG = "FCE4EC"
AMBER_BG = "FFF3E0"

# ── Reusable styles ──
header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=NAVY)
sub_header_font = Font(name="Arial", bold=True, color=DARK_GREY, size=10)
sub_header_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
body_font = Font(name="Arial", size=10, color=DARK_GREY)
bold_font = Font(name="Arial", size=10, color=DARK_GREY, bold=True)
title_font = Font(name="Arial", bold=True, color=NAVY, size=14)
subtitle_font = Font(name="Arial", bold=True, color=MID_BLUE, size=11)
thin_border = Border(
    left=Side(style="thin", color=MID_GREY),
    right=Side(style="thin", color=MID_GREY),
    top=Side(style="thin", color=MID_GREY),
    bottom=Side(style="thin", color=MID_GREY),
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
wrap_align = Alignment(vertical="top", wrap_text=True)

status_fills = {
    "Not Started": PatternFill("solid", fgColor=LIGHT_GREY),
    "In Progress": PatternFill("solid", fgColor=AMBER_BG),
    "Complete": PatternFill("solid", fgColor=GREEN_BG),
    "N/A": PatternFill("solid", fgColor=MID_GREY),
    "Blocked": PatternFill("solid", fgColor=RED_BG),
}

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_sub_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_data_rows(ws, start_row, end_row, max_col, alt=True):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.alignment = wrap_align
            cell.border = thin_border
            if alt and (r - start_row) % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=PALE_BLUE)

def add_title(ws, title, subtitle=None):
    ws.cell(row=1, column=1, value=title).font = title_font
    if subtitle:
        ws.cell(row=2, column=1, value=subtitle).font = subtitle_font
    ws.row_dimensions[1].height = 30
    return 4 if subtitle else 3

def add_data_validation_status(ws, col_letter, start_row, end_row):
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"Not Started,In Progress,Complete,N/A,Blocked"', allow_blank=True)
    dv.error = "Please select a valid status"
    dv.errorTitle = "Invalid Status"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")

def add_priority_validation(ws, col_letter, start_row, end_row):
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")

def add_likelihood_validation(ws, col_letter, start_row, end_row):
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")

# ═══════════════════════════════════════════════════════════════
# TAB 1: OVERVIEW & INSTRUCTIONS
# ═══════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Overview"
ws.sheet_properties.tabColor = NAVY

r = 1
ws.cell(row=r, column=1, value="Public Inquiry Consulting Toolkit").font = Font(name="Arial", bold=True, color=NAVY, size=18)
ws.row_dimensions[r].height = 35
r += 1
ws.cell(row=r, column=1, value="Lifecycle Management Workbook").font = Font(name="Arial", bold=True, color=MID_BLUE, size=13)
r += 2

info = [
    ("Inquiry Name:", ""),
    ("Type:", "Statutory / Non-Statutory"),
    ("Sponsoring Department:", ""),
    ("Chair:", ""),
    ("Secretary:", ""),
    ("Solicitor:", ""),
    ("Counsel:", ""),
    ("Date Established:", ""),
    ("Terms of Reference Date:", ""),
    ("Target Report Date:", ""),
]
for label, val in info:
    ws.cell(row=r, column=1, value=label).font = bold_font
    ws.cell(row=r, column=2, value=val).font = body_font
    ws.cell(row=r, column=2).alignment = left_align
    r += 1

r += 1
ws.cell(row=r, column=1, value="How to use this workbook").font = subtitle_font
r += 1
instructions = [
    "Each tab corresponds to a phase of the inquiry lifecycle. Use the checklists to track progress.",
    "The Decision Log captures key decisions for institutional memory — fill it as you go.",
    "The Risk Register flags common pitfalls. Review it at each phase transition.",
    "The Statutory vs Non-Statutory tab helps frame early scoping conversations.",
    "The Budget Tracker, CP Register, and Stakeholder Map are cross-cutting tools.",
    "Status options: Not Started | In Progress | Complete | N/A | Blocked",
]
for inst in instructions:
    ws.cell(row=r, column=1, value=inst).font = body_font
    ws.cell(row=r, column=1).alignment = wrap_align
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 45
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 20

# ═══════════════════════════════════════════════════════════════
# PHASE TABS — checklist items for each phase
# ═══════════════════════════════════════════════════════════════

phase_data = {
    "1. Establish & Scope": {
        "color": "2C3E6B",
        "subtitle": "From announcement through to finalised terms of reference",
        "items": [
            ("Confirm statutory basis", "Determine whether inquiry will be statutory (Inquiries Act 2005) or non-statutory. Assess need for compulsion powers, public hearing presumption, and legal framework implications.", "High", "Sponsor / Minister"),
            ("Draft and consult on terms of reference", "Terms should set out purpose, matters to investigate, whether recommendations required, reporting line, publication responsibility, and realistic scope. Consult PM, Cabinet Office, devolved administrations, GLD, chair, and affected parties.", "High", "Sponsor / Minister"),
            ("Assess ECHR obligations", "Consider whether Articles 2 or 3 create investigative obligations requiring a public inquiry. Take legal advice.", "High", "GLD / Sponsor"),
            ("Check concurrent proceedings", "Identify any criminal investigations, inquests, regulatory investigations, or civil proceedings that may affect timing, scope, or conduct. Consult Attorney General if needed.", "High", "Sponsor / GLD"),
            ("Scoping exercise", "Officials should examine key issues, likely timescale, cost, volume of evidence, and number of potential witnesses and participants.", "High", "Sponsor Team"),
            ("Consult previous administrations", "If events occurred under a prior government, consult former ministers via Cabinet Secretary before announcement.", "Medium", "Cabinet Secretary"),
            ("Public Sector Equality Duty assessment", "Document how PSED has been considered in decisions about establishing the inquiry and framing terms of reference.", "Medium", "Sponsor"),
            ("Prepare announcement", "Draft ministerial statement including full terms of reference, chair name, panel details, and relevant part of UK. Parliament first when in session.", "High", "Sponsor / Private Office"),
            ("Publish terms of reference", "Finalise and publish. Ensure they are clear, unambiguous, deliverable, and do not extend beyond what is necessary.", "High", "Chair / Sponsor"),
            ("Commission cost and duration estimate", "Provide minister with best assessment of costs, uncertainties, and risks. Reference benchmarks from comparable inquiries.", "Medium", "Sponsor / Secretary"),
        ],
    },
    "2. Appointments & Team": {
        "color": "3A5BA0",
        "subtitle": "Appointing the chair, panel, counsel, secretary and building the team",
        "items": [
            ("Appoint chair", "Identify and appoint chair with appropriate expertise, integrity, leadership, and communication skills. Consult PM for judicial appointments. Consider diversity.", "High", "Minister / Sponsor"),
            ("Assess need for panel members", "Decide whether chair sits alone or with panel. If panel, identify subject matter expertise gaps. Chair must be consulted.", "High", "Minister / Chair"),
            ("Appoint inquiry secretary", "Usually Deputy Director to DG seniority. Key adviser to chair on policy and procedures, responsible for budget and team leadership.", "High", "Sponsor / Chair"),
            ("Appoint solicitor to the inquiry", "Usually from GLD. Main source of legal and procedural advice. Appoint early to avoid procedural issues.", "High", "Chair / GLD"),
            ("Assess need for and appoint counsel", "Required for complex statutory inquiries. Fair, open, non-discriminatory appointment process. Significant cost implications.", "High", "Chair / Solicitor"),
            ("Agree engagement letters and terms", "Cover role, accountability, conflict management, pay, and duration. HM Treasury approval if pay exceeds thresholds.", "High", "Sponsor / Secretary"),
            ("National security vetting", "Consider appropriate vetting level for chair, panel, and key staff based on nature of inquiry and material likely to be handled.", "Medium", "Secretary / Sponsor"),
            ("Conflict of interest checks", "Screen all appointees for conflicts. Document assessment. Consider both actual and perceived conflicts.", "High", "Secretary"),
            ("Build wider secretariat", "Recruit across: subject matter expertise, operations, information management, communications, HR, finance, security. Staff work independently of parent departments.", "Medium", "Secretary"),
            ("Agree staff welfare and support arrangements", "Consider impact of potentially distressing material. Plan trauma-informed training and psychological support from outset.", "Medium", "Secretary"),
            ("Plan for return of staff to parent departments", "Put arrangements in place for Civil Service staff redeployment at inquiry conclusion. Complete reports and appraisals.", "Low", "Secretary / HR"),
        ],
    },
    "3. Infrastructure & Ops": {
        "color": "4A7FB5",
        "subtitle": "Venue, IT, security, accommodation and operational setup",
        "items": [
            ("Identify and secure hearing venue", "Consider: proximity to affected communities, cost, accessibility, public and media capacity, security, separation of participants. Not automatically London.", "High", "Secretary / Sponsor"),
            ("Secure office accommodation", "Sufficient, accessible space with appropriate IT. May co-locate with sponsor department if independence not compromised.", "High", "Secretary / Sponsor"),
            ("Procure IT systems — basic infrastructure", "Laptops, phones, inquiry-branded email, document storage, collaboration tools, access to departmental HR/finance systems.", "High", "Secretary / Sponsor IT"),
            ("Procure eDiscovery / evidence management system", "Required by almost all inquiries for secure evidence storage, review, and management. Do not underestimate procurement timeframes.", "High", "Secretary / Solicitor"),
            ("Set up secure document transfer capability", "For receiving sensitive material from information providers.", "High", "Secretary / IT"),
            ("Commission inquiry website", "Host away from gov.uk. GDS-agreed domain. Content: terms of reference, team bios, procedures, hearing info, transcripts, costs, contact details.", "High", "Secretary / Comms"),
            ("Establish physical security arrangements", "Agree with sponsor department. Cover office, hearing centre, hard copy documents.", "High", "Secretary / DSO"),
            ("Establish data security protocols", "Align with HMG Security Policy Framework. Cover how evidence is held, managed, disclosed, and handle sensitive material.", "High", "Secretary / DPO"),
            ("Agree vetting levels for all staff", "Commensurate with role and inquiry nature. Complete before staff access documents or data.", "High", "Secretary / DSO"),
            ("Procure hearing room broadcast and transcription", "Electronic hearing support, audiovisual broadcast, live transcription services. Consider copyright for AV recordings.", "Medium", "Secretary / Sponsor"),
            ("Engage National Archives early", "For guidance on records management, website preservation, Crown copyright, and archiving planning from the start.", "Medium", "Secretary"),
            ("Register as data controller with ICO", "Inquiry is independent data controller. Appoint DPO, produce privacy notice and data protection policy.", "High", "Secretary / DPO"),
        ],
    },
    "4. Protocols & Procedures": {
        "color": "5B9BD5",
        "subtitle": "Establishing the rules of engagement for the inquiry's work",
        "items": [
            ("Develop issues list from terms of reference", "Led by solicitor and counsel. Treat as living document, kept under review. Share with core participants for proposed additions.", "High", "Solicitor / Counsel"),
            ("Publish provisional timetable", "Include dates for evidence requests, witness statements, oral proceedings, and proposed report publication date. Update regularly.", "High", "Chair / Secretary"),
            ("Draft and publish core participant designation protocol", "Set out criteria, process, and approach. Consider phase-specific designation.", "High", "Chair / Solicitor"),
            ("Draft protocol on legal representation and funding", "Set out approach to public funding of representation. Account for any Section 40 ministerial determination. Include cost controls.", "High", "Chair / Solicitor"),
            ("Request Section 40 determination from minister", "Determines conditions and qualifications on chair's power to award legal costs. Do this shortly after terms of reference finalised.", "High", "Sponsor / Secretary"),
            ("Draft disclosure and document handling protocol", "Cover how inquiry requests, receives, reviews, and discloses documents. Include redaction approach and confidentiality undertakings.", "High", "Solicitor / Counsel"),
            ("Draft witness statement protocol", "Set out process for requesting and preparing statements — whether witness-led, inquiry-led, or hybrid approach.", "High", "Solicitor / Counsel"),
            ("Draft hearing procedure protocol", "Cover questioning of witnesses, role of counsel, opening/closing statements, support for witnesses, breaks, and access arrangements.", "Medium", "Solicitor / Counsel"),
            ("Establish media engagement strategy", "Press office support, approach to broadcasting hearings, transcript publication, media statements at key milestones.", "Medium", "Secretary / Comms"),
            ("Draft restriction order / notice protocol", "Set out approach to restricting attendance, disclosure, or publication. Cover process for applications.", "Medium", "Solicitor"),
            ("Draft redaction protocol", "Process for redacting personal details and irrelevant information from disclosed material. Include representations process for information providers.", "Medium", "Solicitor"),
            ("Agree management statement with sponsor department", "Set out respective roles, responsibilities, and procedures to manage independence/accountability balance.", "Medium", "Secretary / Sponsor"),
            ("Develop internal working practices", "Staff code of conduct, information handling, communication channels, escalation procedures.", "Medium", "Secretary"),
        ],
    },
    "5. Evidence & Investigation": {
        "color": "70AD47",
        "subtitle": "Gathering documentary evidence and witness statements",
        "items": [
            ("Issue written requests for documentary evidence", "Identify information holders. Craft requests carefully — sufficiently broad but targeted. Set deadlines and format requirements.", "High", "Solicitor / Counsel"),
            ("Manage incoming document volumes", "Review for relevance against terms of reference and issues list. Log, index, and store securely in evidence management system.", "High", "Solicitor / Evidence Team"),
            ("Assess need for Section 21 compulsion notices", "For statutory inquiries where informal requests are not complied with, or where providers need formal cover for disclosure.", "Medium", "Chair / Solicitor"),
            ("Handle privilege claims and PII applications", "Take legal advice on claims under Section 22 (legal professional privilege, self-incrimination, parliamentary proceedings). Manage PII balancing exercise.", "Medium", "Solicitor / Counsel"),
            ("Prepare and issue witness statement requests", "Rule 9 requests for statutory inquiries. Develop approach: witness-led, inquiry-led interview, or hybrid. Set timelines.", "High", "Solicitor / Counsel"),
            ("Conduct witness interviews where inquiry-led", "Prepare interview plans. Consider vulnerability, support needs, interpreters. Produce statement for witness approval.", "Medium", "Solicitor / Counsel"),
            ("Disclose relevant material to core participants", "Via document management system. Subject to redactions and confidentiality undertakings. Disclose witness statements before oral evidence.", "High", "Solicitor"),
            ("Manage ongoing disclosure requests and challenges", "Handle disputes about scope, relevance, privilege. Keep disclosure log updated.", "Medium", "Solicitor"),
            ("Commission expert reports or establish expert groups", "Where specialist knowledge needed to understand evidence or support recommendations.", "Medium", "Chair / Counsel"),
            ("Consider innovative evidence-gathering methods", "Seminars, site visits, intermediaries for vulnerable witnesses, listening exercises, pen portraits / commemoration hearings.", "Low", "Chair / Solicitor"),
            ("Conduct National Archives searches", "Use Discovery catalogue. Arrange private access at Kew. Request digital copies as needed.", "Low", "Solicitor / Evidence Team"),
            ("Ongoing review and refinement of issues list", "As evidence emerges, update the issues list. Consult core participants on proposed changes.", "Medium", "Solicitor / Counsel"),
        ],
    },
    "6. Hearings": {
        "color": "BF8F00",
        "subtitle": "Preliminary hearings, oral evidence, and public proceedings",
        "items": [
            ("Plan and hold preliminary hearing(s)", "Set out outline plan, approach to core participants, legal representation, funding, and procedures. Invite evidence from others.", "High", "Chair / Counsel"),
            ("Prepare hearing timetable", "Sequence witnesses logically. Build in breaks, administrative time, and contingency. Publish and share with core participants.", "High", "Counsel / Solicitor"),
            ("Prepare opening statement", "Chair or counsel sets out background, investigative work, issues for oral evidence, procedures, and timescales.", "High", "Counsel / Chair"),
            ("Witness preparation meetings", "Counsel meets witnesses in advance. Explain process, manage expectations, identify support needs.", "High", "Counsel / Solicitor"),
            ("Manage witness support during hearings", "Personal supporters, breaks, psychological support, accessible facilities. Especially for vulnerable witnesses and core participants.", "High", "Secretary / Ops"),
            ("Conduct oral evidence sessions", "Counsel questions witnesses. Manage applications from core participant counsel to ask questions. Chair maintains control.", "High", "Chair / Counsel"),
            ("Manage core participant engagement during hearings", "Handle suggested questions, disclosure of new material, applications for additional witnesses.", "Medium", "Counsel / Solicitor"),
            ("Publish daily transcripts", "Corrected transcripts on inquiry website same day or next morning. Include necessary redactions.", "High", "Secretary / Ops"),
            ("Manage live broadcast of proceedings", "Ensure reliable streaming. Handle any restriction orders requiring closed sessions.", "Medium", "Secretary / IT"),
            ("Handle closed or private hearing sessions", "Where restriction orders or notices require it. Manage separate transcription and record-keeping.", "Medium", "Chair / Solicitor"),
            ("Receive closing statements from core participants", "Set aside time after oral evidence. Provides opportunity for observations and suggested recommendations.", "Medium", "Chair / Counsel"),
            ("Manage media throughout hearing period", "Press statements, briefings, managing public interest. Maintain balance with sub judice concerns.", "Medium", "Secretary / Comms"),
            ("Monitor for judicial review risk", "Track procedural decisions that could be challenged. Document reasoning. 14-day challenge window.", "Medium", "Solicitor"),
        ],
    },
    "7. Report & Closure": {
        "color": "C00000",
        "subtitle": "Writing the report, Maxwellisation, publication, and closing down",
        "items": [
            ("Agree report writing approach", "Who drafts which sections — counsel, solicitor, chair, or combination. Consider engaging editor or copy-editor for style consistency.", "High", "Chair"),
            ("Draft report", "Must address terms of reference, be supported by evidence, use clear language, include executive summary and recommendations.", "High", "Chair / Counsel"),
            ("Conduct Maxwellisation / warning letter process", "Send warning letters to anyone who may be subject of explicit or significant criticism. Allow reasonable time for representations.", "High", "Chair / Solicitor"),
            ("Reviews and checks before publication", "Full review for personal data, protected information, accuracy of evidence references, typographical errors, and escaped criticisms.", "High", "Solicitor / Editor"),
            ("Agree publication responsibility and process", "Confirm whether minister or chair publishes. Agree practical steps including sensitivity checking by sponsor department.", "High", "Chair / Sponsor"),
            ("Manage advance access for minister", "Balance minister's need to prepare parliamentary response against perception of independence and victims' expectations.", "High", "Chair / Secretary"),
            ("Organise lock-in for core participants", "Venue, security, separate rooms if needed, device surrender, confidentiality undertakings, staggered access periods.", "High", "Secretary / Ops"),
            ("Arrange laying before Parliament", "Coordinate with parliamentary authorities. Prepare written or oral ministerial statement. Arrange opposition leader access.", "High", "Sponsor / Secretary"),
            ("Publish report", "Website publication, chair's public statement, print run for key recipients. Coordinate timing with parliamentary laying.", "High", "Chair / Secretary"),
            ("Prepare and submit lessons learned paper", "Secretary writes within two months of inquiry end. Cover timetable, costs, accommodation, IT, sponsor relationship, difficulties, good practice.", "High", "Secretary"),
            ("Terminate contracts and vacate premises", "Hearing space, offices, IT equipment, phone lines, email accounts, utilities. Allow buffer period for unexpected applications.", "Medium", "Secretary / Ops"),
            ("Archive and transfer records to National Archives", "Index all documents. Destroy duplicates methodically with destruction record. Transfer to TNA or sponsor department as directed.", "High", "Secretary / TNA"),
            ("Communicate inquiry closure to stakeholders", "Advance notice of when phone lines and email will cease. Direct future queries to sponsor department.", "Medium", "Secretary / Comms"),
            ("Transition witness and stakeholder support", "Agree with sponsor department what support continues, in what form, and who funds it.", "Medium", "Secretary / Sponsor"),
            ("Monitor recommendation implementation", "Consider chair's ongoing role. Government should respond within six months. Annual updates to Parliament until closed.", "Medium", "Chair / Sponsor"),
        ],
    },
}

for tab_name, data in phase_data.items():
    ws = wb.create_sheet(title=tab_name)
    ws.sheet_properties.tabColor = data["color"]

    start = add_title(ws, tab_name, data["subtitle"])

    headers = ["#", "Action Item", "Description / Guidance", "Priority", "Responsible Role", "Status", "Target Date", "Notes"]
    col_widths = [5, 35, 55, 10, 20, 14, 14, 30]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=start, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header_row(ws, start, len(headers))

    data_start = start + 1
    for idx, (action, desc, priority, role) in enumerate(data["items"], 1):
        r = data_start + idx - 1
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=action)
        ws.cell(row=r, column=3, value=desc)
        ws.cell(row=r, column=4, value=priority)
        ws.cell(row=r, column=5, value=role)
        ws.cell(row=r, column=6, value="Not Started")
        ws.cell(row=r, column=7, value="")
        ws.cell(row=r, column=8, value="")

    data_end = data_start + len(data["items"]) - 1
    style_data_rows(ws, data_start, data_end, len(headers))
    add_data_validation_status(ws, "F", data_start, data_end)
    add_priority_validation(ws, "D", data_start, data_end)

    ws.auto_filter.ref = f"A{start}:{get_column_letter(len(headers))}{data_end}"
    ws.freeze_panes = f"A{data_start}"

# ═══════════════════════════════════════════════════════════════
# DECISION LOG
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet(title="Decision Log")
ws.sheet_properties.tabColor = "7030A0"

start = add_title(ws, "Decision Log", "Capturing key decisions for institutional memory")

headers = ["#", "Date", "Phase", "Decision", "Options Considered", "Rationale", "Decided By", "Implications / Dependencies", "Review Date"]
col_widths = [5, 12, 18, 35, 35, 35, 18, 30, 12]
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    ws.cell(row=start, column=i, value=h)
    ws.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws, start, len(headers))

from openpyxl.worksheet.datavalidation import DataValidation
phase_dv = DataValidation(type="list",
    formula1='"1. Establish & Scope,2. Appointments & Team,3. Infrastructure & Ops,4. Protocols & Procedures,5. Evidence & Investigation,6. Hearings,7. Report & Closure,Cross-cutting"',
    allow_blank=True)
ws.add_data_validation(phase_dv)

for r in range(start + 1, start + 51):
    ws.cell(row=r, column=1, value=r - start)
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).font = body_font
        ws.cell(row=r, column=c).alignment = wrap_align
        ws.cell(row=r, column=c).border = thin_border
        if (r - start) % 2 == 0:
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=PALE_BLUE)
phase_dv.add(f"C{start+1}:C{start+50}")
ws.auto_filter.ref = f"A{start}:{get_column_letter(len(headers))}{start+50}"
ws.freeze_panes = f"A{start+1}"

# ═══════════════════════════════════════════════════════════════
# RISK REGISTER
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet(title="Risk Register")
ws.sheet_properties.tabColor = "C00000"

start = add_title(ws, "Risk Register", "Common pitfalls and risks across the inquiry lifecycle")

headers = ["#", "Phase", "Risk Description", "Likelihood", "Impact", "Risk Rating", "Mitigation", "Owner", "Status", "Review Date"]
col_widths = [5, 18, 40, 12, 12, 12, 40, 18, 14, 12]
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    ws.cell(row=start, column=i, value=h)
    ws.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws, start, len(headers))

risks = [
    ("1. Establish & Scope", "Terms of reference too broad or ambiguous, leading to scope creep, cost overruns, and delay", "Medium", "High", "Ensure ToR are clear, unambiguous, and deliverable. Consult widely. Include explicit exclusions.", "Chair / Sponsor"),
    ("1. Establish & Scope", "Failure to identify concurrent criminal proceedings, causing prejudice or requiring delay", "Medium", "High", "Conduct thorough check with CPS, police, and Attorney General before establishment.", "Sponsor / GLD"),
    ("2. Appointments & Team", "Chair appointment challenged on grounds of bias or conflict of interest", "Low", "High", "Thorough conflict screening. Document assessment. Consider judicial review risk in selection.", "Sponsor"),
    ("2. Appointments & Team", "Difficulty recruiting experienced secretary or solicitor, delaying start", "Medium", "Medium", "Begin recruitment early. Consider secondments from other inquiries. Cabinet Office can advise.", "Sponsor"),
    ("3. Infrastructure & Ops", "IT procurement delays leaving inquiry without evidence management system", "Medium", "High", "Do not underestimate procurement timeframes. Have plan ready for incoming chair. Consider framework agreements.", "Secretary / Sponsor"),
    ("3. Infrastructure & Ops", "Data breach of sensitive evidence material", "Low", "High", "Robust data security from day one. Align with HMG Security Policy Framework. Vetting before access.", "Secretary / DPO"),
    ("4. Protocols & Procedures", "Core participants excluded from protocol development, leading to challenge or loss of cooperation", "Medium", "Medium", "Consult core participants on draft protocols. Allow reasonable time for representations.", "Chair / Solicitor"),
    ("4. Protocols & Procedures", "Failure to make Section 40 determination early, causing funding disputes", "Medium", "Medium", "Request determination shortly after ToR finalised. Publish costs protocol early.", "Sponsor / Secretary"),
    ("5. Evidence & Investigation", "Information providers fail to cooperate or delay disclosure", "Medium", "High", "Escalate from informal to formal requests. Use Section 21 compulsion powers. Set clear deadlines.", "Solicitor / Chair"),
    ("5. Evidence & Investigation", "Overwhelmed by volume of disclosed material", "High", "Medium", "Target requests carefully. Use eDiscovery tools. Prioritise review by relevance to issues list.", "Solicitor / Evidence Team"),
    ("6. Hearings", "Judicial review challenge to procedural decision causing delay", "Medium", "High", "Document reasoning for all procedural decisions. Monitor 14-day challenge window. Budget for potential JR costs.", "Solicitor"),
    ("6. Hearings", "Inadequate witness support leading to poor evidence or reputational damage", "Medium", "Medium", "Trauma-informed approach. Psychological support available. Personal supporters. Accessible facilities.", "Secretary / Ops"),
    ("7. Report & Closure", "Maxwellisation process takes longer than planned, delaying publication", "High", "Medium", "Build sufficient time into timetable from outset. Set clear deadlines for representations.", "Chair / Solicitor"),
    ("7. Report & Closure", "Minister seeks extended advance access, undermining perception of independence", "Medium", "Medium", "Agree advance access arrangements early. Limit to preparation of parliamentary response. Inquiry team present during review.", "Chair / Secretary"),
    ("7. Report & Closure", "Records not properly archived, creating future FOI and accountability problems", "Medium", "Medium", "Engage National Archives from start. Plan records management throughout, not just at closure. Index all destroyed documents.", "Secretary / TNA"),
    ("Cross-cutting", "Budget overruns without adequate financial controls", "High", "High", "Preliminary budget agreed early. Regular monitoring. Sponsor manages delegation per Managing Public Money principles.", "Secretary / Sponsor"),
    ("Cross-cutting", "Loss of public confidence due to perceived delays or lack of transparency", "Medium", "High", "Publish provisional timetable and updates. Regular cost publication. Proactive communications strategy.", "Chair / Secretary"),
    ("Cross-cutting", "Staff burnout from distressing material and high-pressure environment", "High", "Medium", "Welfare support from outset. Trauma-informed training. Regular check-ins. Access to counselling.", "Secretary / HR"),
]

data_start = start + 1
for idx, (phase, desc, like, impact, mitigation, owner) in enumerate(risks, 1):
    r = data_start + idx - 1
    ws.cell(row=r, column=1, value=idx)
    ws.cell(row=r, column=2, value=phase)
    ws.cell(row=r, column=3, value=desc)
    ws.cell(row=r, column=4, value=like)
    ws.cell(row=r, column=5, value=impact)
    rating_cell = get_column_letter(6)
    ws.cell(row=r, column=6).value = f'=IF(AND(D{r}="High",E{r}="High"),"Critical",IF(OR(D{r}="High",E{r}="High"),"High",IF(AND(D{r}="Low",E{r}="Low"),"Low","Medium")))'
    ws.cell(row=r, column=7, value=mitigation)
    ws.cell(row=r, column=8, value=owner)
    ws.cell(row=r, column=9, value="Open")
    ws.cell(row=r, column=10, value="")

data_end = data_start + len(risks) - 1
style_data_rows(ws, data_start, data_end, len(headers))
add_likelihood_validation(ws, "D", data_start, data_end + 20)
add_likelihood_validation(ws, "E", data_start, data_end + 20)

status_dv = DataValidation(type="list", formula1='"Open,Mitigating,Closed,Accepted"', allow_blank=True)
ws.add_data_validation(status_dv)
status_dv.add(f"I{data_start}:I{data_end+20}")

ws.auto_filter.ref = f"A{start}:{get_column_letter(len(headers))}{data_end}"
ws.freeze_panes = f"A{data_start}"

# ═══════════════════════════════════════════════════════════════
# STATUTORY VS NON-STATUTORY MATRIX
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet(title="Statutory vs Non-Statutory")
ws.sheet_properties.tabColor = "548235"

start = add_title(ws, "Statutory vs Non-Statutory Decision Matrix", "Key differences to inform scoping advice")

headers = ["Dimension", "Statutory (Inquiries Act 2005)", "Non-Statutory", "Consulting Considerations"]
col_widths = [25, 40, 40, 40]
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    ws.cell(row=start, column=i, value=h)
    ws.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws, start, len(headers))

matrix = [
    ("Legal framework", "Governed by Inquiries Act 2005 and Inquiry Rules 2006. Codified powers and procedures.", "No binding legal framework. Chair determines procedure within terms of reference.", "Statutory provides certainty but less flexibility. Non-statutory allows innovation but carries risk of challenge on fairness grounds."),
    ("Power to compel evidence", "Chair can compel witnesses to attend, give evidence under oath, and produce documents (s.21). Criminal sanctions for non-compliance.", "No power to compel. Relies on voluntary cooperation. Minister should seek assurances from information providers.", "If cooperation is uncertain — particularly from reluctant organisations — statutory basis is strongly advisable."),
    ("Core participants", "Formal designation under Inquiry Rules with specific rights: opening/closing statements, advance disclosure, questioning through counsel.", "No formal concept. Chair may grant equivalent rights but no statutory basis. Sponsor may choose to fund representation.", "Core participant framework provides structure for managing multiple parties. Without it, managing participation requires careful protocol design."),
    ("Public hearings", "Rebuttable presumption of public hearings. Must do what is reasonable to ensure public access.", "No presumption. May be held largely in private if terms of reference allow.", "Public confidence often requires public hearings. If inquiry can operate effectively in private, non-statutory may be faster and cheaper."),
    ("Immunity", "Statutory immunity from civil action for inquiry personnel. Parliamentary privilege for reports.", "No statutory immunity. Sponsor department should provide indemnity in writing. Report may need parliamentary privilege via Return to an Address.", "Immunity is significant protection. Without it, inquiry personnel face greater personal risk. Indemnity arrangements must be robust."),
    ("Warning letters", "Mandatory under Inquiry Rules if report contains explicit or significant criticism. Formal Maxwellisation process.", "No statutory requirement, but fairness demands a similar process. Many non-statutory inquiries adopt equivalent procedures.", "Either way, budget time for Maxwellisation. Non-statutory inquiries that skip it risk challenge on fairness grounds."),
    ("Publication", "Minister responsible but can delegate to chair. Must lay before Parliament. Minister may withhold material on specified grounds.", "Minister responsible unless delegated to chair. Should be laid before Parliament. Consider Return to an Address for parliamentary privilege.", "Publication process is similar in practice. The key difference is the statutory protection for the report content."),
    ("FOI", "Not a public authority during lifetime — exempt from FOI requests. FOI applies after records deposited.", "Also not a public authority — exempt during lifetime. Same post-closure position.", "No practical difference during inquiry lifetime. Post-closure FOI obligations apply equally."),
    ("Judicial review", "Decisions subject to judicial review. 14-day time limit from awareness of decision.", "Decisions also subject to judicial review on same basis. Same 14-day time limit.", "Both are equally vulnerable to challenge. Statutory framework may actually reduce JR risk by providing clear procedural basis."),
    ("Cost and duration", "Often more expensive and longer. Average ~3 years for completed statutory inquiries since 2000.", "Generally cheaper and faster. Average ~2 years. But varies widely — some have exceeded statutory inquiries in duration.", "Cost saving is not guaranteed. A complex non-statutory inquiry can be just as expensive. The real driver is scope, not form."),
    ("Conversion", "N/A — already statutory.", "Can be converted to statutory under the Inquiries Act if cooperation fails or compulsion powers become necessary.", "Always consider whether non-statutory is viable first. Conversion is available as a fallback but causes disruption and delay."),
    ("Data protection", "Independent data controller. Must register with ICO, appoint DPO, produce privacy notice.", "Independent data controller. Same GDPR and Data Protection Act 2018 obligations.", "No practical difference. Both require full data protection compliance from the outset."),
]

data_start = start + 1
for idx, (dim, stat, nonstat, consult) in enumerate(matrix):
    r = data_start + idx
    ws.cell(row=r, column=1, value=dim)
    ws.cell(row=r, column=2, value=stat)
    ws.cell(row=r, column=3, value=nonstat)
    ws.cell(row=r, column=4, value=consult)
data_end = data_start + len(matrix)
style_data_rows(ws, data_start, data_end, len(headers))
ws.freeze_panes = f"A{data_start}"

# ═══════════════════════════════════════════════════════════════
# BUDGET TRACKER
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet(title="Budget Tracker")
ws.sheet_properties.tabColor = "BF8F00"

start = add_title(ws, "Budget Tracker", "Monitoring inquiry expenditure against budget")

headers = ["Cost Category", "Budget (£)", "Spend to Date (£)", "Committed (£)", "Forecast Total (£)", "Variance (£)", "Variance %", "Notes"]
col_widths = [30, 15, 15, 15, 15, 15, 12, 30]
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    ws.cell(row=start, column=i, value=h)
    ws.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws, start, len(headers))

categories = [
    "Chair and panel fees",
    "Counsel fees",
    "Solicitor and legal team costs",
    "Core participant legal costs",
    "Secretariat staff salaries",
    "Accommodation — office",
    "Accommodation — hearing venue",
    "IT systems and evidence management",
    "Website hosting and development",
    "Transcription services",
    "Broadcast and AV",
    "Witness expenses",
    "Witness and staff welfare support",
    "Travel and subsistence",
    "Security",
    "Communications and media",
    "Expert and assessor fees",
    "Printing and publication",
    "Archiving and records management",
    "Other / contingency",
]

data_start = start + 1
for idx, cat in enumerate(categories):
    r = data_start + idx
    ws.cell(row=r, column=1, value=cat)
    for c in [2, 3, 4, 5]:
        ws.cell(row=r, column=c).number_format = '#,##0'
    ws.cell(row=r, column=6).value = f'=B{r}-E{r}'
    ws.cell(row=r, column=6).number_format = '#,##0'
    ws.cell(row=r, column=7).value = f'=IF(B{r}=0,"-",F{r}/B{r})'
    ws.cell(row=r, column=7).number_format = '0.0%'

data_end = data_start + len(categories) - 1
style_data_rows(ws, data_start, data_end, len(headers), alt=True)

# Totals row
total_r = data_end + 1
ws.cell(row=total_r, column=1, value="TOTAL").font = Font(name="Arial", bold=True, color=WHITE, size=10)
ws.cell(row=total_r, column=1).fill = PatternFill("solid", fgColor=NAVY)
for c in [2, 3, 4, 5, 6]:
    col_l = get_column_letter(c)
    ws.cell(row=total_r, column=c).value = f'=SUM({col_l}{data_start}:{col_l}{data_end})'
    ws.cell(row=total_r, column=c).font = Font(name="Arial", bold=True, color=WHITE, size=10)
    ws.cell(row=total_r, column=c).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row=total_r, column=c).number_format = '#,##0'
    ws.cell(row=total_r, column=c).border = thin_border
ws.cell(row=total_r, column=7).value = f'=IF(B{total_r}=0,"-",F{total_r}/B{total_r})'
ws.cell(row=total_r, column=7).font = Font(name="Arial", bold=True, color=WHITE, size=10)
ws.cell(row=total_r, column=7).fill = PatternFill("solid", fgColor=NAVY)
ws.cell(row=total_r, column=7).number_format = '0.0%'
ws.cell(row=total_r, column=7).border = thin_border
ws.cell(row=total_r, column=8).fill = PatternFill("solid", fgColor=NAVY)
ws.cell(row=total_r, column=8).border = thin_border

ws.freeze_panes = f"A{data_start}"

# ═══════════════════════════════════════════════════════════════
# CORE PARTICIPANT REGISTER
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet(title="CP Register")
ws.sheet_properties.tabColor = "7030A0"

start = add_title(ws, "Core Participant Register", "Tracking core participants, legal representation, and funding")

headers = ["#", "Name / Organisation", "Type", "Phase(s) Designated", "Date Designated", "Recognised Legal Rep", "Funding Status", "Joint Representation Group", "Key Contact", "Notes"]
col_widths = [5, 25, 15, 18, 12, 25, 15, 20, 20, 25]
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    ws.cell(row=start, column=i, value=h)
    ws.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws, start, len(headers))

type_dv = DataValidation(type="list", formula1='"Individual,Organisation,Government Body,Action Group,Other"', allow_blank=True)
ws.add_data_validation(type_dv)

funding_dv = DataValidation(type="list", formula1='"Public Funded,Self-Funded,Application Pending,Not Applicable"', allow_blank=True)
ws.add_data_validation(funding_dv)

for r in range(start + 1, start + 51):
    ws.cell(row=r, column=1, value=r - start)
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).font = body_font
        ws.cell(row=r, column=c).alignment = wrap_align
        ws.cell(row=r, column=c).border = thin_border
        if (r - start) % 2 == 0:
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=PALE_BLUE)
type_dv.add(f"C{start+1}:C{start+50}")
funding_dv.add(f"G{start+1}:G{start+50}")
ws.freeze_panes = f"A{start+1}"

# ═══════════════════════════════════════════════════════════════
# STAKEHOLDER MAP
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet(title="Stakeholder Map")
ws.sheet_properties.tabColor = "548235"

start = add_title(ws, "Stakeholder Map", "Key relationships and engagement approach")

headers = ["#", "Stakeholder", "Category", "Interest / Role", "Influence", "Engagement Level", "Key Contact", "Engagement Approach", "Notes"]
col_widths = [5, 25, 18, 30, 12, 15, 20, 30, 25]
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    ws.cell(row=start, column=i, value=h)
    ws.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws, start, len(headers))

cat_dv = DataValidation(type="list",
    formula1='"Sponsor Department,Chair / Panel,Inquiry Team,Core Participant,Witness,Victims / Families,Media,Government Body,NGO / Campaign Group,Expert / Assessor,Legal Representative,Other"',
    allow_blank=True)
ws.add_data_validation(cat_dv)

influence_dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
ws.add_data_validation(influence_dv)

engage_dv = DataValidation(type="list", formula1='"Manage Closely,Keep Satisfied,Keep Informed,Monitor"', allow_blank=True)
ws.add_data_validation(engage_dv)

for r in range(start + 1, start + 51):
    ws.cell(row=r, column=1, value=r - start)
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).font = body_font
        ws.cell(row=r, column=c).alignment = wrap_align
        ws.cell(row=r, column=c).border = thin_border
        if (r - start) % 2 == 0:
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=PALE_BLUE)
cat_dv.add(f"C{start+1}:C{start+50}")
influence_dv.add(f"E{start+1}:E{start+50}")
engage_dv.add(f"F{start+1}:F{start+50}")
ws.freeze_panes = f"A{start+1}"

# ═══════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════
output_path = "/sessions/lucid-youthful-gates/mnt/practical guide/Inquiry Consulting Toolkit.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
